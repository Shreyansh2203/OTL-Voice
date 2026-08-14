
from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from .connection import db_path, get_conn
from .passwords import verify_password
from .seed import load_seed

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:  # pragma: no cover - dependency guidance
    raise SystemExit(
        "This export needs openpyxl, which is not installed.\n"
        "    pip install openpyxl"
    ) from exc

DEFAULT_OUTPUT = "otl_reference.xlsx"

# Passwords are stored as PBKDF2 hashes. They are not reversible, but a full
# 100-character hash per row makes the sheet unreadable and there is no reason to
# scatter them through a file people email around — show the parameters only.
_TRUNCATE = {"password_hash"}
_TRUNCATE_AT = 28

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT = Font(italic=True, color="808080")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FONT = Font(bold=True, color="9C5700")
MONO_FONT = Font(name="Consolas")
OK_FONT = Font(color="1E7B34")
BAD_FONT = Font(bold=True, color="C00000")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Sheet name -> the query behind it. Ordered parent-first so the workbook reads
# in the same direction as the foreign keys.
TABLE_SHEETS: List[tuple[str, str, str]] = [
    (
        "employees",
        "SELECT employee_id, username, full_name, is_active, password_hash, created_at"
        " FROM employees ORDER BY employee_id",
        "Front-end sign-in. employee_id is written to OTL as Employee_Number_c.",
    ),
    (
        "work_orders",
        "SELECT work_order, description, is_active FROM work_orders ORDER BY work_order",
        "OTL Work_Order_c. One work order holds many projects.",
    ),
    (
        "projects",
        "SELECT project_no, work_order, project_name, is_active FROM projects"
        " ORDER BY work_order, project_no",
        "OTL Project_No_c / Project_Name_c. work_order is a single NOT NULL foreign "
        "key, which is what makes 'one work order, many projects -- never the "
        "reverse' true by construction.",
    ),
    (
        "project_tasks",
        "SELECT task_id, project_no, task_details, is_active FROM project_tasks"
        " ORDER BY project_no, task_id",
        "OTL Tasks_Details_c. Capped at 80 characters by a CHECK constraint.",
    ),
    (
        "assignments",
        "SELECT employee_id, project_no, assigned_at FROM assignments"
        " ORDER BY employee_id, project_no",
        "Who may log against what. Deliberately holds no work_order or "
        "project_name column: both are determined by project_no, so storing them "
        "here would let them drift out of sync with the projects table.",
    ),
    (
        "v_employee_labour",
        "SELECT employee_id, username, full_name, work_order, project_no,"
        " project_name, task_id, task_details FROM v_employee_labour"
        " ORDER BY employee_id, work_order, project_no, task_id",
        "The flat join -- assignments with their work order and project name "
        "filled in. This is the shape to read; the columns are derived, not stored.",
    ),
]


# --------------------------------------------------------------------------- #
# Sheet helpers
# --------------------------------------------------------------------------- #
def _cell_value(column: str, value: Any) -> Any:
    if value is None:
        return ""
    if column in _TRUNCATE:
        text = str(value)
        return text if len(text) <= _TRUNCATE_AT else text[:_TRUNCATE_AT] + "…"
    return value


def _write_note(sheet: Worksheet, row: int, note: str) -> int:
    cell = sheet.cell(row=row, column=1, value=note)
    cell.font = NOTE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    sheet.row_dimensions[row].height = 30
    return row + 2


def _write_table(
    sheet: Worksheet,
    columns: Sequence[str],
    rows: Sequence[Sequence[Any]],
    start_row: int,
) -> None:
    for index, name in enumerate(columns, start=1):
        cell = sheet.cell(row=start_row, column=index, value=name)
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, BORDER

    for offset, record in enumerate(rows, start=1):
        for index, (name, value) in enumerate(zip(columns, record), start=1):
            cell = sheet.cell(
                row=start_row + offset, column=index, value=_cell_value(name, value)
            )
            cell.border = BORDER

    _autosize(sheet, columns, rows, start_row)
    # Freeze the header so long tables stay readable while scrolling.
    sheet.freeze_panes = sheet.cell(row=start_row + 1, column=1)
    if rows:
        sheet.auto_filter.ref = (
            f"A{start_row}:{get_column_letter(len(columns))}{start_row + len(rows)}"
        )


def _autosize(
    sheet: Worksheet,
    columns: Sequence[str],
    rows: Sequence[Sequence[Any]],
    start_row: int,
) -> None:
    for index, name in enumerate(columns, start=1):
        widest = len(str(name))
        for record in rows:
            widest = max(widest, len(str(_cell_value(name, record[index - 1]))))
        sheet.column_dimensions[get_column_letter(index)].width = min(widest + 3, 60)


def _add_table_sheet(
    book: Workbook, conn: sqlite3.Connection, name: str, query: str, note: str
) -> int:
    sheet = book.create_sheet(title=name[:31])
    cursor = conn.execute(query)
    columns = [d[0] for d in cursor.description]
    rows = [tuple(r) for r in cursor.fetchall()]

    title = sheet.cell(row=1, column=1, value=name)
    title.font = TITLE_FONT
    _write_table(sheet, columns, rows, _write_note(sheet, 2, note))
    return len(rows)


# --------------------------------------------------------------------------- #
# Derived sheets
# --------------------------------------------------------------------------- #
def _add_catalogue_sheet(book: Workbook, conn: sqlite3.Connection) -> None:
    sheet = book.create_sheet(title="Catalogue tree")
    title = sheet.cell(row=1, column=1, value="Labour catalogue")
    title.font = TITLE_FONT
    row = _write_note(
        sheet,
        2,
        "Work order -> project -> task. A project appears under exactly one work "
        "order; its name and tasks follow from its project number.",
    )

    header = ["Work order", "Project no.", "Project name", "Task details"]
    rows: List[List[Any]] = []
    for order in conn.execute(
        "SELECT work_order, description FROM work_orders ORDER BY work_order"
    ).fetchall():
        rows.append([order["work_order"], "", order["description"] or "", ""])
        for project in conn.execute(
            "SELECT project_no, project_name FROM projects WHERE work_order = ?"
            " ORDER BY project_no",
            (order["work_order"],),
        ).fetchall():
            rows.append(["", project["project_no"], project["project_name"], ""])
            for task in conn.execute(
                "SELECT task_details FROM project_tasks WHERE project_no = ?"
                " ORDER BY task_id",
                (project["project_no"],),
            ).fetchall():
                rows.append(["", "", "", task["task_details"]])

    _write_table(sheet, header, rows, row)
    # Bold each work-order line so the grouping is visible at a glance.
    for offset, record in enumerate(rows, start=1):
        if record[0]:
            for column in range(1, len(header) + 1):
                sheet.cell(row=row + offset, column=column).font = Font(bold=True)


def _add_matrix_sheet(book: Workbook, conn: sqlite3.Connection) -> None:
    sheet = book.create_sheet(title="Assignment matrix")
    title = sheet.cell(row=1, column=1, value="Who works on what")
    title.font = TITLE_FONT
    row = _write_note(
        sheet,
        2,
        "X marks an assignment. Column headers show the project number with its "
        "work order in brackets.",
    )

    employees = conn.execute(
        "SELECT employee_id, full_name FROM employees ORDER BY employee_id"
    ).fetchall()
    projects = conn.execute(
        "SELECT project_no, project_name, work_order FROM projects"
        " ORDER BY work_order, project_no"
    ).fetchall()
    assigned = {
        (r["employee_id"], r["project_no"])
        for r in conn.execute("SELECT employee_id, project_no FROM assignments")
    }

    header = ["Employee", "Name"] + [
        f"{p['project_no']} - {p['project_name']} (WO {p['work_order']})"
        for p in projects
    ]
    rows = [
        [e["employee_id"], e["full_name"]]
        + ["X" if (e["employee_id"], p["project_no"]) in assigned else "" for p in projects]
        for e in employees
    ]

    _write_table(sheet, header, rows, row)
    for index in range(3, len(header) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 16
        sheet.cell(row=row, column=index).alignment = Alignment(
            wrap_text=True, vertical="bottom"
        )
    sheet.row_dimensions[row].height = 46


def _add_credentials_sheet(book: Workbook, conn: sqlite3.Connection) -> None:
    # index=0 here and for the Overview, which is added straight after, so the
    # workbook opens on Overview with the credentials sheet next to it.
    sheet = book.create_sheet(title="Login credentials", index=0)
    title = sheet.cell(row=1, column=1, value="Demo sign-in credentials")
    title.font = TITLE_FONT

    banner = sheet.cell(
        row=2,
        column=1,
        value=(
            "DUMMY ACCOUNTS — local test data only. Passwords are read from "
            "backend/db/seed.json; the database itself stores only PBKDF2 hashes. "
            "Do not reuse these for anything real, and do not treat this sheet as "
            "a credential store."
        ),
    )
    banner.fill, banner.font = WARN_FILL, WARN_FONT
    banner.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    sheet.row_dimensions[2].height = 44

    seeded = {
        str(e["employee_id"]): e.get("password", "")
        for e in load_seed().get("employees", [])
    }
    employees = conn.execute(
        "SELECT employee_id, username, full_name, is_active, password_hash"
        " FROM employees ORDER BY employee_id"
    ).fetchall()

    header = ["Username", "Password", "Employee ID", "Full name", "Active", "Verified"]
    rows: List[List[Any]] = []
    for employee in employees:
        password = seeded.get(employee["employee_id"])
        if password is None:
            shown, verdict = "(not in seed.json)", "unknown"
        elif verify_password(password, employee["password_hash"]):
            shown, verdict = password, "yes — signs in"
        else:
            shown, verdict = password, "NO — hash differs"
        rows.append(
            [
                employee["username"],
                shown,
                employee["employee_id"],
                employee["full_name"],
                "yes" if employee["is_active"] else "no",
                verdict,
            ]
        )

    start = 4
    _write_table(sheet, header, rows, start)
    for offset, record in enumerate(rows, start=1):
        sheet.cell(row=start + offset, column=2).font = MONO_FONT
        verdict_cell = sheet.cell(row=start + offset, column=6)
        verdict_cell.font = OK_FONT if record[5].startswith("yes") else BAD_FONT

    # Employees present in seed.json but missing from the database — usually a
    # sign the DB predates a seed.json edit and needs `seed --force`.
    missing = sorted(set(seeded) - {e["employee_id"] for e in employees})
    if missing:
        note = sheet.cell(
            row=start + len(rows) + 2,
            column=1,
            value=(
                "In seed.json but not in the database: "
                + ", ".join(missing)
                + " — run: python -m backend.db.seed --force"
            ),
        )
        note.font = BAD_FONT

    for column, width in zip("ABCDEF", (18, 16, 13, 22, 9, 18)):
        sheet.column_dimensions[column].width = width


def _add_overview_sheet(book: Workbook, counts: Dict[str, int], source: Path) -> None:
    sheet = book.create_sheet(title="Overview", index=0)
    title = sheet.cell(row=1, column=1, value="OTL reference tables")
    title.font = Font(bold=True, size=15)

    row = _write_note(sheet, 2, f"Exported from {source}")
    rows: List[tuple] = [
        (name, counts.get(name, 0), note) for name, _query, note in TABLE_SHEETS
    ]
    rows += [
        ("Catalogue tree", "", "Derived: work order -> project -> tasks, indented."),
        ("Assignment matrix", "", "Derived: employees x projects, X where assigned."),
        (
            "Login credentials",
            counts.get("employees", 0),
            "Derived: demo usernames and passwords (from seed.json), each checked "
            "against the stored hash. Dummy accounts only.",
        ),
    ]
    _write_table(sheet, ["Sheet", "Rows", "What it holds"], rows, row)

    sheet.column_dimensions["C"].width = 90
    for offset in range(1, len(rows) + 1):
        sheet.cell(row=row + offset, column=3).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        sheet.row_dimensions[row + offset].height = 44


# --------------------------------------------------------------------------- #
# Entry point
# --------------------------------------------------------------------------- #
def export(output: Optional[Path] = None) -> Path:
    source = db_path()
    if not source.is_file():
        raise SystemExit(
            f"No reference database at {source}.\n"
            "    python -m backend.db.seed"
        )

    target = Path(output) if output else source.parent / DEFAULT_OUTPUT
    target.parent.mkdir(parents=True, exist_ok=True)

    book = Workbook()
    book.remove(book.active)  # drop the default sheet openpyxl creates

    with get_conn() as conn:
        counts = {
            name: _add_table_sheet(book, conn, name, query, note)
            for name, query, note in TABLE_SHEETS
        }
        _add_catalogue_sheet(book, conn)
        _add_matrix_sheet(book, conn)
        _add_credentials_sheet(book, conn)
        _add_overview_sheet(book, counts, source)

    _check_writable(target)
    book.save(target)
    return target


def _check_writable(target: Path) -> None:
    if not target.exists():
        return
    try:
        with open(target, "r+b"):
            pass
    except PermissionError as exc:
        raise SystemExit(
            f"Could not write {target} — it is open in Excel (or otherwise "
            f"locked).\nClose it and run again, or pass -o to write elsewhere."
        ) from exc


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export the OTL reference tables to an Excel workbook."
    )
    parser.add_argument(
        "-o", "--output", type=Path, default=None,
        help=f"destination .xlsx (default: alongside the database, {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()

    target = export(args.output)
    print(f"Wrote {target}")


if __name__ == "__main__":
    main()
