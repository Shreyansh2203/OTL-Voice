import base64
import datetime
import json
import os
import sys
import urllib.parse
import xml.etree.ElementTree as ET
from pathlib import Path

# Ensure UTF-8 output in Windows terminals
if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

project_root = Path(__file__).resolve().parent
env_path = project_root / ".env"
data_dir = project_root / "data"
data_dir.mkdir(parents=True, exist_ok=True)

try:
    from dotenv import load_dotenv
    load_dotenv(dotenv_path=env_path)
except ImportError:
    pass

import httpx

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

USERNAME = os.getenv("OTL_SERVICE_USERNAME", "").strip()
PASSWORD = os.getenv("OTL_SERVICE_PASSWORD", "")
DEFAULT_OTL_URL = os.getenv(
    "OTL_BASE_URL",
    "https://fa-epxp-test-saasfaprod1.fa.ocs.oraclecloud.com/hcmRestApi/resources/11.13.18.05/timeRecordEventRequests"
)

if not USERNAME or not PASSWORD:
    print(f"[ERROR] OTL_SERVICE_USERNAME or OTL_SERVICE_PASSWORD missing in {env_path}")
    sys.exit(1)

parsed = urllib.parse.urlparse(DEFAULT_OTL_URL)
HOST_URL = f"{parsed.scheme}://{parsed.netloc}"

client = httpx.Client(
    auth=(USERNAME, PASSWORD),
    timeout=60.0,
    headers={"Accept": "application/json", "Accept-Encoding": "gzip"},
)

print("=" * 70)
print(" Oracle Fusion Master Data Extraction")
print(f" Target Host: {HOST_URL}")
print(f" User:        {USERNAME}")
print("=" * 70)

master_data = {
    "metadata": {
        "extracted_at": datetime.datetime.now(datetime.UTC).isoformat(),
        "host": HOST_URL,
        "authenticated_user": USERNAME,
    },
    "summary": {},
    "employees": [],
    "projects": [],
    "project_tasks": [],
    "project_allocations": [],
    "timecards": [],
}

# --------------------------------------------------------------------------- #
# 1. Fetch Employees (via BIP Report for fast bulk retrieval)
# --------------------------------------------------------------------------- #
print("\n[1/5] Extracting Employees via Oracle BI Publisher...")
bip_url = f"{HOST_URL}/xmlpserver/services/v2/ReportService"
report_path = "/Custom/EmployeeSyncReport.xdo"
soap_body = f"""<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" xmlns:v2="http://xmlns.oracle.com/oxp/service/v2">
   <soapenv:Header/>
   <soapenv:Body>
      <v2:runReport>
         <v2:reportRequest>
            <v2:attributeFormat>xml</v2:attributeFormat>
            <v2:attributeLocale>en-US</v2:attributeLocale>
            <v2:reportAbsolutePath>{report_path}</v2:reportAbsolutePath>
         </v2:reportRequest>
         <v2:userID>{USERNAME}</v2:userID>
         <v2:password>{PASSWORD}</v2:password>
      </v2:runReport>
   </soapenv:Body>
</soapenv:Envelope>"""

try:
    resp = client.post(bip_url, content=soap_body.encode("utf-8"), headers={"Content-Type": "text/xml;charset=UTF-8", "SOAPAction": '""'})
    if resp.status_code == 200:
        root = ET.fromstring(resp.text)
        report_bytes = None
        for elem in root.iter():
            if elem.tag.endswith("reportBytes") or elem.tag == "reportBytes":
                report_bytes = elem
                break
        if report_bytes is not None and report_bytes.text:
            raw_xml = base64.b64decode(report_bytes.text).decode("utf-8", errors="replace")
            data_root = ET.fromstring(raw_xml)
            for row in data_root.findall(".//G_1"):
                emp_no = (row.findtext("EMPLOYEE_NUMBER") or "").strip()
                person_id = (row.findtext("HCM_PERSON_ID") or "").strip()
                name = (row.findtext("EMPLOYEE_NAME") or "").strip()
                created = (row.findtext("CREATED_BY") or "").strip()
                master_data["employees"].append({
                    "employee_number": emp_no,
                    "person_id": person_id,
                    "full_name": name,
                    "created_by": created,
                })
            print(f"  -> Extracted {len(master_data['employees'])} employees.")
    else:
        print(f"  -> BIP returned {resp.status_code}, falling back to REST /workers...")
except Exception as e:
    print(f"  -> BIP call error: {e}, falling back to REST...")

# Fallback to REST /workers if BIP failed or empty
if not master_data["employees"]:
    print("  -> Fetching employees via REST /workers...")
    w_resp = client.get(f"{HOST_URL}/hcmRestApi/resources/11.13.18.05/workers", params={"expand": "names", "limit": 100})
    if w_resp.status_code == 200:
        for w in w_resp.json().get("items", []):
            names = w.get("names", [])
            if isinstance(names, dict):
                names = names.get("items", [])
            name = names[0].get("DisplayName") if names else "N/A"
            master_data["employees"].append({
                "employee_number": str(w.get("PersonNumber") or ""),
                "person_id": str(w.get("PersonId") or ""),
                "full_name": name,
                "created_by": str(w.get("CreatedBy") or ""),
            })
        print(f"  -> Extracted {len(master_data['employees'])} employees via REST.")

# --------------------------------------------------------------------------- #
# 2. Fetch Projects (PPM REST API)
# --------------------------------------------------------------------------- #
print("\n[2/5] Extracting Projects via PPM REST API...")
proj_resp = client.get(
    f"{HOST_URL}/fscmRestApi/resources/11.13.18.05/projects",
    params={"fields": "ProjectId,ProjectNumber,ProjectName,ProjectManagerName,ProjectStatus,ProjectStartDate,ProjectEndDate,ProjectDescription,BusinessUnitName", "limit": 100}
)

if proj_resp.status_code == 200:
    projects_raw = proj_resp.json().get("items", [])
    for p in projects_raw:
        proj_obj = {
            "project_id": str(p.get("ProjectId") or ""),
            "project_number": str(p.get("ProjectNumber") or ""),
            "project_name": p.get("ProjectName") or "",
            "status": p.get("ProjectStatus") or "",
            "manager": p.get("ProjectManagerName") or "",
            "start_date": p.get("ProjectStartDate") or "",
            "end_date": p.get("ProjectEndDate") or "",
            "description": p.get("ProjectDescription") or "",
            "business_unit": p.get("BusinessUnitName") or "",
            "tasks": [],
            "allocations": [],
        }
        master_data["projects"].append(proj_obj)
    print(f"  -> Extracted {len(master_data['projects'])} projects.")
else:
    print(f"  -> Projects query failed ({proj_resp.status_code}): {proj_resp.text[:200]}")

# --------------------------------------------------------------------------- #
# 3 & 4. Fetch Project Tasks & Allocations (Child Endpoints)
# --------------------------------------------------------------------------- #
print("\n[3/5 & 4/5] Extracting Project Tasks & Team Allocations for each Project...")
for proj in master_data["projects"]:
    p_id = proj["project_id"]
    p_num = proj["project_number"]
    p_name = proj["project_name"]

    # 3. Tasks
    task_resp = client.get(f"{HOST_URL}/fscmRestApi/resources/11.13.18.05/projects/{p_id}/child/Tasks", params={"limit": 50})
    if task_resp.status_code == 200:
        tasks = task_resp.json().get("items", [])
        for t in tasks:
            t_obj = {
                "project_number": p_num,
                "project_name": p_name,
                "project_id": p_id,
                "task_id": str(t.get("TaskId") or ""),
                "task_number": str(t.get("TaskNumber") or ""),
                "task_name": t.get("TaskName") or "",
                "description": t.get("Description") or "",
                "start_date": t.get("StartDate") or "",
                "completion_date": t.get("CompletionDate") or "",
            }
            proj["tasks"].append(t_obj)
            master_data["project_tasks"].append(t_obj)

    # 4. Allocations / Team Members
    tm_resp = client.get(f"{HOST_URL}/fscmRestApi/resources/11.13.18.05/projects/{p_id}/child/ProjectTeamMembers", params={"limit": 50})
    if tm_resp.status_code == 200:
        members = tm_resp.json().get("items", [])
        for m in members:
            m_obj = {
                "project_number": p_num,
                "project_name": p_name,
                "project_id": p_id,
                "person_name": m.get("PersonName") or "",
                "project_role": m.get("ProjectRole") or "",
                "email": m.get("Email") or "",
                "start_date": m.get("StartDate") or "",
                "end_date": m.get("EndDate") or "",
            }
            proj["allocations"].append(m_obj)
            master_data["project_allocations"].append(m_obj)

print(f"  -> Extracted {len(master_data['project_tasks'])} total project tasks.")
print(f"  -> Extracted {len(master_data['project_allocations'])} total project allocations.")

# --------------------------------------------------------------------------- #
# 5. Fetch Timecards (OTL REST API)
# --------------------------------------------------------------------------- #
print("\n[5/5] Extracting Timecard Events via OTL REST API...")
otl_resp = client.get(
    f"{HOST_URL}/hcmRestApi/resources/11.13.18.05/timeRecordEventRequests",
    params={"limit": 50}
)
if otl_resp.status_code == 200:
    for tc in otl_resp.json().get("items", []):
        tc_id = str(tc.get("timeRecordEventRequestId") or "")
        
        # Try fetching child events for detailed attributes
        details = []
        events_resp = client.get(f"{HOST_URL}/hcmRestApi/resources/11.13.18.05/timeRecordEventRequests/{tc_id}/child/timeRecordEvent")
        if events_resp.status_code == 200:
            for ev in events_resp.json().get("items", []):
                attrs = ev.get("timeRecordEventAttribute", [])
                comment = next((a["attributeValue"] for a in attrs if a.get("attributeName") == "Comment"), "")
                payroll = next((a["attributeValue"] for a in attrs if a.get("attributeName") == "PayrollTimeType"), "")
                details.append({
                    "reporter_id": ev.get("reporterId") or "",
                    "reporter_type": ev.get("reporterIdType") or "",
                    "measure_hours": ev.get("measure") or 0,
                    "start_time": ev.get("startTime") or "",
                    "stop_time": ev.get("stopTime") or "",
                    "operation_type": ev.get("operationType") or "",
                    "comment": comment,
                    "payroll_type": payroll,
                })
        
        master_data["timecards"].append({
            "request_id": tc_id,
            "process_mode": tc.get("processMode") or "",
            "process_inline": tc.get("processInline") or "",
            "events_count": len(details),
            "events": details,
        })
    print(f"  -> Extracted {len(master_data['timecards'])} timecard request records.")
else:
    print(f"  -> Timecards query failed ({otl_resp.status_code}): {otl_resp.text[:200]}")

# Summary Stats
master_data["summary"] = {
    "total_employees": len(master_data["employees"]),
    "total_projects": len(master_data["projects"]),
    "total_project_tasks": len(master_data["project_tasks"]),
    "total_project_allocations": len(master_data["project_allocations"]),
    "total_timecards": len(master_data["timecards"]),
}

# --------------------------------------------------------------------------- #
# Output 1: Structured JSON
# --------------------------------------------------------------------------- #
json_out_path = data_dir / "fusion_master_catalogue.json"
with open(json_out_path, "w", encoding="utf-8") as f:
    json.dump(master_data, f, indent=2, ensure_ascii=False)

# --------------------------------------------------------------------------- #
# Output 2: Structured Multi-Sheet Excel (.xlsx)
# --------------------------------------------------------------------------- #
xlsx_out_path = data_dir / "fusion_master_catalogue.xlsx"
if HAVE_OPENPYXL:
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    def style_sheet(ws, headers, rows):
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for r in rows:
            ws.append(r)

        # Autofit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Sheet 1: Summary
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum.append(["Entity Name", "Record Count", "Description"])
    for cell in ws_sum[1]:
        cell.font = header_font
        cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    ws_sum.append(["Employees / Workers", len(master_data["employees"]), "Active Oracle HCM Employees"])
    ws_sum.append(["Projects", len(master_data["projects"]), "PPM Project Definitions"])
    ws_sum.append(["Project Tasks", len(master_data["project_tasks"]), "Granular tasks under projects"])
    ws_sum.append(["Project Allocations", len(master_data["project_allocations"]), "Team members / assigned personnel"])
    ws_sum.append(["Timecards", len(master_data["timecards"]), "OTL Timecard Submission Events"])
    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 40

    # Sheet 2: Employees
    ws_emp = wb.create_sheet(title="Employees")
    emp_headers = ["Employee Number", "Person ID", "Full Name", "Created By"]
    emp_rows = [[e["employee_number"], e["person_id"], e["full_name"], e["created_by"]] for e in master_data["employees"]]
    style_sheet(ws_emp, emp_headers, emp_rows)

    # Sheet 3: Projects
    ws_proj = wb.create_sheet(title="Projects")
    proj_headers = ["Project Number", "Project Name", "Status", "Manager", "Business Unit", "Start Date", "End Date", "Project ID"]
    proj_rows = [[p["project_number"], p["project_name"], p["status"], p["manager"], p["business_unit"], p["start_date"], p["end_date"], p["project_id"]] for p in master_data["projects"]]
    style_sheet(ws_proj, proj_headers, proj_rows)

    # Sheet 4: Project Tasks
    ws_tasks = wb.create_sheet(title="Project Tasks")
    task_headers = ["Project Number", "Project Name", "Task Number", "Task Name", "Task ID", "Start Date", "Completion Date"]
    task_rows = [[t["project_number"], t["project_name"], t["task_number"], t["task_name"], t["task_id"], t["start_date"], t["completion_date"]] for t in master_data["project_tasks"]]
    style_sheet(ws_tasks, task_headers, task_rows)

    # Sheet 5: Project Allocations
    ws_alloc = wb.create_sheet(title="Project Allocations")
    alloc_headers = ["Project Number", "Project Name", "Person Name", "Project Role", "Email", "Start Date", "End Date"]
    alloc_rows = [[a["project_number"], a["project_name"], a["person_name"], a["project_role"], a["email"], a["start_date"], a["end_date"]] for a in master_data["project_allocations"]]
    style_sheet(ws_alloc, alloc_headers, alloc_rows)

    # Sheet 6: Timecards
    ws_tc = wb.create_sheet(title="Timecards")
    tc_headers = ["Request ID", "Process Mode", "Inline", "Reporter ID", "Hours", "Comment", "Start Time", "Stop Time", "Payroll Type"]
    tc_rows = []
    for tc in master_data["timecards"]:
        if tc["events"]:
            for ev in tc["events"]:
                tc_rows.append([
                    tc["request_id"], tc["process_mode"], tc["process_inline"],
                    ev["reporter_id"], ev["measure_hours"], ev["comment"],
                    ev["start_time"], ev["stop_time"], ev["payroll_type"]
                ])
        else:
            tc_rows.append([tc["request_id"], tc["process_mode"], tc["process_inline"], "", "", "", "", "", ""])
    style_sheet(ws_tc, tc_headers, tc_rows)

    wb.save(xlsx_out_path)

print("\n" + "=" * 70)
print(" EXTRACTION & EXPORT COMPLETE!")
print("=" * 70)
print(" Summary of Extracted Data:")
print(f"  * Employees:           {len(master_data['employees'])}")
print(f"  * Projects:            {len(master_data['projects'])}")
print(f"  * Project Tasks:       {len(master_data['project_tasks'])}")
print(f"  * Project Allocations: {len(master_data['project_allocations'])}")
print(f"  * Timecards:           {len(master_data['timecards'])}")
print("-" * 70)
print(" Structured Master Files Created:")
print(f"  1. Excel (.xlsx): {xlsx_out_path}")
print(f"  2. JSON  (.json): {json_out_path}")
print("=" * 70)
