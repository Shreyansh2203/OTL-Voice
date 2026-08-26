import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

data_dir = Path(__file__).resolve().parent / "data"
catalogue_path = data_dir / "fusion_master_catalogue.json"
if not catalogue_path.is_file():
    print(f"[ERROR] {catalogue_path} not found. Run export_fusion_master.py first.")
    import sys
    sys.exit(1)
with open(catalogue_path, "r", encoding="utf-8") as f:
    master_data = json.load(f)
employees = master_data.get("employees", [])
projects = master_data.get("projects", [])
project_tasks = master_data.get("project_tasks", [])
project_allocations = master_data.get("project_allocations", [])
timecards = master_data.get("timecards", [])
projects_by_num = {}
for p in projects:
    p_num = str(p.get("project_number") or "").strip()
    if p_num:
        projects_by_num[p_num] = p
tasks_by_proj = {}
for t in project_tasks:
    p_num = str(t.get("project_number") or "").strip()
    if p_num not in tasks_by_proj:
        tasks_by_proj[p_num] = []
    tasks_by_proj[p_num].append(t)
allocations_by_name = {}
for a in project_allocations:
    name = (a.get("person_name") or "").strip().lower()
    if name:
        if name not in allocations_by_name:
            allocations_by_name[name] = []
        allocations_by_name[name].append(a)
timecards_by_emp = {}
for tc in timecards:
    for ev in tc.get("events", []):
        rep_id = str(ev.get("reporter_id") or "").strip()
        if rep_id:
            if rep_id not in timecards_by_emp:
                timecards_by_emp[rep_id] = []
            timecards_by_emp[rep_id].append({
                "request_id": tc.get("request_id"),
                "process_mode": tc.get("process_mode"),
                "hours": ev.get("measure_hours"),
                "comment": ev.get("comment"),
                "start_time": ev.get("start_time"),
                "stop_time": ev.get("stop_time"),
                "payroll_type": ev.get("payroll_type"),
            })
person_master_list = []
flat_rows = []
for emp in employees:
    emp_no = emp.get("employee_number") or ""
    person_id = emp.get("person_id") or ""
    full_name = emp.get("full_name") or ""
    created_by = emp.get("created_by") or ""
    matched_allocs = allocations_by_name.get(full_name.lower(), [])
    assigned_projects = []
    seen_proj_nums = set()
    for a in matched_allocs:
        p_num = str(a.get("project_number") or "")
        if p_num and p_num not in seen_proj_nums:
            seen_proj_nums.add(p_num)
            p_obj = projects_by_num.get(p_num, {})
            assigned_projects.append({
                "project_number": p_num,
                "project_name": a.get("project_name") or p_obj.get("project_name") or "",
                "role": a.get("project_role") or "Team Member",
                "manager": p_obj.get("manager") or "",
                "status": p_obj.get("status") or "Active",
                "tasks": [
                    {
                        "task_number": t.get("task_number"),
                        "task_name": t.get("task_name"),
                        "task_id": t.get("task_id")
                    } for t in tasks_by_proj.get(p_num, [])
                ]
            })
    for p_num, p_obj in projects_by_num.items():
        mgr = (p_obj.get("manager") or "").strip().lower()
        if mgr and (mgr in full_name.lower() or full_name.lower() in mgr) and p_num not in seen_proj_nums:
            seen_proj_nums.add(p_num)
            assigned_projects.append({
                "project_number": p_num,
                "project_name": p_obj.get("project_name") or "",
                "role": "Project Manager",
                "manager": p_obj.get("manager") or full_name,
                "status": p_obj.get("status") or "Active",
                "tasks": [
                    {
                        "task_number": t.get("task_number"),
                        "task_name": t.get("task_name"),
                        "task_id": t.get("task_id")
                    } for t in tasks_by_proj.get(p_num, [])
                ]
            })
    emp_timecards = timecards_by_emp.get(emp_no, [])
    person_entry = {
        "employee_number": emp_no,
        "person_id": person_id,
        "full_name": full_name,
        "created_by": created_by,
        "total_assigned_projects": len(assigned_projects),
        "total_logged_timecards": len(emp_timecards),
        "assigned_projects": assigned_projects,
        "timecards": emp_timecards,
    }
    person_master_list.append(person_entry)
    if assigned_projects:
        for p in assigned_projects:
            tasks_str = "; ".join([f"#{t['task_number']} {t['task_name']}" for t in p["tasks"]]) if p["tasks"] else "No tasks assigned"
            tc_summary = f"{len(emp_timecards)} timecard(s)" if emp_timecards else "None"
            flat_rows.append([
                emp_no, person_id, full_name, created_by,
                p["project_number"], p["project_name"], p["role"], p["status"],
                tasks_str, tc_summary
            ])
    else:
        tc_summary = f"{len(emp_timecards)} timecard(s)" if emp_timecards else "None"
        flat_rows.append([
            emp_no, person_id, full_name, created_by,
            "Not Assigned", "N/A", "N/A", "N/A",
            "N/A", tc_summary
        ])
json_out = data_dir / "fusion_person_master.json"
with open(json_out, "w", encoding="utf-8") as f:
    json.dump({
        "metadata": master_data.get("metadata", {}),
        "total_persons": len(person_master_list),
        "persons": person_master_list
    }, f, indent=2, ensure_ascii=False)
xlsx_out = data_dir / "fusion_person_master.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Person Master Directory"
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
headers = [
    "Employee #", "Person ID", "Full Name", "Created By",
    "Assigned Project #", "Project Name", "Project Role", "Project Status",
    "Available Tasks", "Timecard Submissions"
]
ws.append(headers)
for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
for r in flat_rows:
    ws.append(r)
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 50)
wb.save(xlsx_out)
print("=" * 70)
print(" UNIFIED PERSON-CENTRIC MASTER FILE CREATED!")
print("=" * 70)
print(f" Total Employees Processed: {len(person_master_list)}")
print(f" Total Flattened Rows:      {len(flat_rows)}")
print("-" * 70)
print(" Files Generated:")
print(f"  1. Excel (.xlsx): {xlsx_out}")
print(f"  2. JSON  (.json): {json_out}")
print("=" * 70)